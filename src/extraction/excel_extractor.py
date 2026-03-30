from openpyxl import load_workbook
from datetime import datetime
import logging

def extract_excel(file_path):
    results = []

    try:
        workbook = load_workbook(file_path)
        logging.info(f"Opened Excel file: {file_path}")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_data = []

            for row in sheet.iter_rows(values_only=True):
                rows_data.append(list(row))

            results.append({
                "content": rows_data,
                "metadata": {
                    "file_name": file_path,
                    "sheet_name": sheet_name,
                    "document_type": "excel",
                    "extraction_date": datetime.utcnow()
                }
            })

    except Exception as e:
        logging.error(f"Error processing Excel file {file_path}: {e}")

    return results